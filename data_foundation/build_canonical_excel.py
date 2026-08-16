from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from .forensic_crawler import SourceRecord, canonical_dataset_sha, tails27, EXPECTED

LABELS = ("DB", "G1", "G2", "G3", "G4", "G5", "G6", "G7")

def split_prizes(full):
    out, pos = [], 0
    for n in EXPECTED:
        out.append(" ".join(full[pos:pos+n]))
        pos += n
    return out

def build_xlsx(records: list[SourceRecord], output: str | Path, conflicts=None):
    output = Path(output); output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook(); ws = wb.active; ws.title = "CanonicalFull27"
    headers = ["Ngày", *LABELS, "TAIL_27", "Source_Count", "Sources", "Record_SHA256", "Dataset_SHA256"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="D9EAF7")
    dataset_sha = canonical_dataset_sha(records)
    for r in sorted(records, key=lambda x: x.draw_date, reverse=True):
        source_names = r.source_domain.removeprefix("QUORUM:").split(";") if r.source_domain.startswith("QUORUM:") else [r.source_domain]
        ws.append([r.draw_date, *split_prizes(r.full_prizes), " ".join(tails27(r.full_prizes)), len([x for x in source_names if x]), ";".join(source_names), r.full_fingerprint, dataset_sha])
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in col) + 2, 70)
        ws.column_dimensions[col[0].column_letter].width = width
    meta = wb.create_sheet("Manifest")
    for row in [("key","value"),("schema","CANONICAL_FULL27_XLSX_V1"),("record_count",len(records)),("dataset_sha256",dataset_sha),("quorum_conflict_dates",len(conflicts or {})),("information_policy","FULL_PRIZES_AUTHORITATIVE; TAIL_27_DERIVED"),("tail_policy","last two digits of each full prize"),("legacy_reference","Ket_Qua_Loto27.xlsx is derived tail-only data")]:
        meta.append(row)
    meta["A1"].font = Font(bold=True); meta["B1"].font = Font(bold=True)
    wb.save(output)
    return dataset_sha
