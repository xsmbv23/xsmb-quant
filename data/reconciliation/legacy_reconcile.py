from __future__ import annotations
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from openpyxl import load_workbook
import json, hashlib

EXPECTED_TAILS = 27

@dataclass(frozen=True)
class LegacyRow:
    day: date
    tails: tuple[str, ...]

def parse_day(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, str):
        for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
            try: return datetime.strptime(v.strip(), fmt).date()
            except ValueError: pass
    raise ValueError(f'unsupported date: {v!r}')

def load_legacy(path: str | Path) -> list[LegacyRow]:
    ws = load_workbook(path, read_only=True, data_only=True).active
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0]: continue
        tails = tuple(str(r[1]).split())
        if len(tails) != EXPECTED_TAILS:
            raise ValueError(f'{r[0]}: expected 27 tails, got {len(tails)}')
        if any(len(x) != 2 or not x.isdigit() for x in tails):
            raise ValueError(f'{r[0]}: invalid tail value')
        rows.append(LegacyRow(parse_day(r[0]), tails))
    if len({r.day for r in rows}) != len(rows):
        raise ValueError('duplicate dates')
    return sorted(rows, key=lambda r: r.day)

def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()

def build_calendar_ledger(rows: list[LegacyRow]) -> list[dict]:
    days = {r.day for r in rows}
    start, end = min(days), max(days)
    out = []
    d = start
    while d <= end:
        if d in days:
            state = 'LEGACY_PRESENT_TAIL_ONLY'
            reason = 'LEGACY_EXCEL_ROW_PRESENT_BUT_FULL_27_UNAVAILABLE'
            evidence = 'Ket_Qua_Loto27.xlsx'
        else:
            state = 'UNKNOWN_GAP'
            reason = 'NO_LEGACY_ROW; AUTHORITATIVE_NON_DRAW_EVIDENCE_REQUIRED'
            evidence = None
        out.append({'date': d.isoformat(), 'state': state, 'reason': reason, 'evidence': evidence})
        d += timedelta(days=1)
    return out

def report(path: str | Path) -> dict:
    rows = load_legacy(path)
    ledger = build_calendar_ledger(rows)
    gaps = [x for x in ledger if x['state'] == 'UNKNOWN_GAP']
    return {
        'schema_version': 'LEGACY_RECONCILIATION_V1',
        'source_file': Path(path).name,
        'row_count': len(rows),
        'oldest': rows[0].day.isoformat(),
        'newest': rows[-1].day.isoformat(),
        'all_rows_have_27_tails': True,
        'unknown_gap_days': len(gaps),
        'canonical_full27': False,
        'promotion': 'DENY',
        'calendar_ledger': ledger,
    }

if __name__ == '__main__':
    import sys
    print(json.dumps(report(sys.argv[1]), ensure_ascii=False, indent=2))
