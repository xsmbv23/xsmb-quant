from __future__ import annotations
from datetime import timedelta
from pathlib import Path
from openpyxl import load_workbook
from data.reconciliation.legacy_reconcile import parse_day
import json

EXPECTED_TAILS = 27

def build(path: str | Path):
    ws = load_workbook(path, read_only=True, data_only=True).active
    present = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0]: continue
        d = parse_day(r[0]); tails = str(r[1]).split()
        if len(tails) != EXPECTED_TAILS:
            raise ValueError(f'{d}: not 27 tails')
        present.append(d)
    days = set(present); start, end = min(days), max(days)
    out = []; d = start
    while d <= end:
        if d in days:
            out.append({'date': d.isoformat(), 'state': 'LEGACY_PRESENT_TAIL_ONLY', 'evidence': 'Ket_Qua_Loto27.xlsx'})
        else:
            out.append({'date': d.isoformat(), 'state': 'UNKNOWN_GAP', 'evidence_required': True})
        d += timedelta(days=1)
    return {'schema_version': 'CALENDAR_LEDGER_V1', 'start': start.isoformat(), 'end': end.isoformat(), 'rows': out, 'promotion': 'DENY'}

if __name__ == '__main__':
    import sys
    print(json.dumps(build(sys.argv[1]), ensure_ascii=False, indent=2))
